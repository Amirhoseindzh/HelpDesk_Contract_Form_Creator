from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from settings.config import PERSIAN_FONT
from repositories.pcform_repo import PCFormRepository
from exports.pdf_converter import form_docx_to_pdf_handler
from utils.widget_utils import set_icon


# constants
DATA_RETRIEVE_ERROR = "Could not retrieve record data."


class DatabaseInfo(ctk.CTkFrame):
    """database info frame with treeview
    and action buttons for export, view, favorite, print, copy"""

    def __init__(self, parent, parent_window):
        super().__init__(parent)
        self.parent_window = parent_window
        self.selected_record_id = None
        self.sort_column = "id"
        self.sort_ascending = True
        # Repository instance
        self.db = PCFormRepository()

        # Create button frame for actions (split into multiple rows for more buttons)
        self.button_frame = ctk.CTkFrame(self)
        self.button_frame.pack(fill="x", padx=5, pady=5)

        # Row 1: Main actions
        self.export_button = ctk.CTkButton(
            self.button_frame,
            text="📄 Export Contract",
            command=self.export_selected_contract,
            width=150,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=40,
            fg_color="green",
            hover_color="darkgreen",
            state="disabled",
        )
        self.export_button.pack(side="left", padx=5, pady=5)

        self.view_details_button = ctk.CTkButton(
            self.button_frame,
            text="👁️ View Details",
            command=self.view_selected_details,
            width=120,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=40,
            state="disabled",
        )
        self.view_details_button.pack(side="left", padx=5, pady=5)

        # Star/Favorite button
        self.favorite_button = ctk.CTkButton(
            self.button_frame,
            text="⭐ Mark Favorite",
            command=self.toggle_favorite,
            width=130,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=40,
            fg_color="orange",
            state="disabled",
        )
        self.favorite_button.pack(side="left", padx=5, pady=5)

        # Print button
        self.print_button = ctk.CTkButton(
            self.button_frame,
            text="🖨️ Print",
            command=self.print_record,
            width=100,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=40,
            state="disabled",
        )
        self.print_button.pack(side="left", padx=5, pady=5)

        # Copy button
        self.copy_button = ctk.CTkButton(
            self.button_frame,
            text="📋 Copy",
            command=self.copy_record,
            width=100,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=40,
            state="disabled",
        )
        self.copy_button.pack(side="left", padx=5, pady=5)

        # Excel export button
        try:
            self.excel_button = ctk.CTkButton(
                self.button_frame,
                text="📊 Export Excel",
                command=self.export_to_excel,
                width=130,
                font=ctk.CTkFont(family=PERSIAN_FONT, size=12, weight="bold"),
                height=40,
                fg_color="darkgreen",
            )
            self.excel_button.pack(side="left", padx=5, pady=5)
        except ImportError as e:
            print(f"openpyxl not available, Excel export disabled: {str(e)}")
            # CSV export fallback (always available)

        self.csv_button = ctk.CTkButton(
            self.button_frame,
            text="📥 Export CSV",
            command=self.export_to_csv,
            width=120,
            font=ctk.CTkFont(family=PERSIAN_FONT, size=12, weight="bold"),
            height=40,
            fg_color="#2d7bdc",
        )
        self.csv_button.pack(side="left", padx=5, pady=5)

        self.refresh_button = ctk.CTkButton(
            self.button_frame,
            text="🔄 Refresh",
            command=self.populate_treeview,
            width=100,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=40,
        )
        self.refresh_button.pack(side="right", padx=5, pady=5)

        # Create scrollable frame for treeview
        self.tree_frame = ctk.CTkFrame(self)
        self.tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Use repository's valid columns as the title list (include id)
        self.title_list = PCFormRepository.VALID_COLUMNS.copy()
        if "is_favorite" not in self.title_list:
            self.title_list.append("is_favorite")

        # In-memory rows and filter state
        self.rows = []
        self.filtered_rows = []
        self._filters_active = False

        # Adjust columns as needed
        self.tree = ttk.Treeview(
            self.tree_frame,
            columns=self.title_list,
            show="headings",
            selectmode="browse",
        )

        # Configure column headings with larger font - make sortable
        style = ttk.Style()
        style.configure("Treeview", font=(PERSIAN_FONT, 12), rowheight=30)
        style.configure(
            "Treeview.Heading", font=(PERSIAN_FONT, 13, "bold"), rowheight=35
        )
        for column in self.title_list:
            self.tree.heading(
                column, text=column, command=lambda col=column: self.sort_by_column(col)
            )
            width = 50 if column == "is_favorite" else 160
            self.tree.column(column, width=width, anchor="w")

        # Add scrollbars
        scrollbar_y = ttk.Scrollbar(
            self.tree_frame, orient="vertical", command=self.tree.yview
        )
        scrollbar_x = ttk.Scrollbar(
            self.tree_frame, orient="horizontal", command=self.tree.xview
        )
        self.tree.configure(
            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set
        )

        # Pack treeview and scrollbars
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

        # Bind selection event
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Double-1>", lambda e: self.export_selected_contract())

        # Populate Treeview with data
        self.populate_treeview()

    def on_select(self, _event):
        """Handle row selection"""
        # _event parameter required by tkinter bind
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            # ID is stored as the first value in the values tuple
            values = item.get("values", [])
            self.selected_record_id = values[0] if values else None
            # Enable buttons
            self.export_button.configure(state="normal")
            self.view_details_button.configure(state="normal")
            self.favorite_button.configure(state="normal")
            self.print_button.configure(state="normal")
            self.copy_button.configure(state="normal")
        else:
            # Disable buttons if nothing selected
            self.export_button.configure(state="disabled")
            self.view_details_button.configure(state="disabled")
            self.favorite_button.configure(state="disabled")
            self.print_button.configure(state="disabled")
            self.copy_button.configure(state="disabled")
            self.selected_record_id = None

    def export_to_csv(self):
        """Export current treeview data to CSV (fallback if openpyxl missing)"""
        import csv

        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile="contracts_export.csv",
            )
            if not file_path:
                return

            with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)
                # headers
                writer.writerow(self.title_list)
                for item in self.tree.get_children():
                    item_data = self.tree.item(item)
                    row = list(item_data.get("values", []))
                    writer.writerow(row)

            messagebox.showinfo("Success", f"Exported to CSV:\n{file_path}")
        except (IOError, OSError, ValueError) as e:
            messagebox.showerror("Error", f"CSV export failed:\n{str(e)}")

    def sort_by_column(self, column):
        if self.sort_column == column:
            self.sort_ascending = not self.sort_ascending
        else:
            self.sort_column = column
            self.sort_ascending = True

        def sort_key(r):
            value = r.get(column)
            return (value is None, str(value).lower())

        self.filtered_rows.sort(
            key=sort_key,
            reverse=not self.sort_ascending
        )

        self._render_tree(self.filtered_rows)


    def toggle_favorite(self):
        """Toggle favorite status for selected record"""
        if not self.selected_record_id:
            messagebox.showwarning("No Selection", "Please select a record.")
            return

        try:
            new_status = self.db.toggle_favorite(int(self.selected_record_id))
        except Exception:
            new_status = None
        if new_status is not None:
            messagebox.showinfo(
                "Success",
                f"Record marked as {'favorite ⭐' if new_status else 'not favorite'}",
            )
            self.populate_treeview()
        else:
            messagebox.showerror("Error", "Failed to update favorite status.")

    def print_record(self):
        """Print selected record"""
        if not self.selected_record_id:
            messagebox.showwarning("No Selection", "Please select a record to print.")
            return

        try:
            record = self.db.get_by_id(int(self.selected_record_id))
        except Exception:
            record = None
        if not record:
            messagebox.showerror("Error", DATA_RETRIEVE_ERROR)
            return

        # Create a formatted text version for printing
        print_text = "=" * 60 + "\n"
        print_text += "COMPUTER SERVICE CONTRACT - PRINT VIEW\n"
        print_text += "=" * 60 + "\n\n"

        for key, value in record.items():
            if key != "id":
                formatted_key = key.replace("_", " ").title()
                print_text += f"{formatted_key}: {value}\n"

        # Open print dialog
        try:
            # Try to print using Notepad print function
            temp_file = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt")],
                initialfile=f"print_{record.get('fullname', 'record')}.txt",
            )
            if temp_file:
                with open(temp_file, "w", encoding="utf-8") as f:
                    f.write(print_text)
                messagebox.showinfo(
                    "Success", f"Record ready to print. File saved to:\n{temp_file}"
                )
        except (IOError, OSError) as e:
            messagebox.showerror("Error", f"Print failed: {str(e)}")

    def copy_record(self):
        """Copy record data to clipboard"""
        if not self.selected_record_id:
            messagebox.showwarning("No Selection", "Please select a record to copy.")
            return

        try:
            record = self.db.get_by_id(int(self.selected_record_id))
        except Exception:
            record = None
        if not record:
            messagebox.showerror("Error", DATA_RETRIEVE_ERROR)
            return

        # Format record as text
        copy_text = ""
        for key, value in record.items():
            if key != "id":
                formatted_key = key.replace("_", " ").title()
                copy_text += f"{formatted_key}: {value}\n"

        # Copy to clipboard using tkinter
        self.clipboard_clear()
        self.clipboard_append(copy_text)
        messagebox.showinfo("Success", "Record copied to clipboard!")

    def export_to_excel(self):
        """Export all records or filtered results to Excel"""

        try:
            # Get current data from treeview
            wb = Workbook()
            ws = wb.active
            ws.title = "Contracts"

            # Add headers with styling
            self._add_excel_headers(ws)

            # Add data from treeview
            self._add_excel_data(ws)

            # Adjust column widths
            self._adjust_excel_columns(ws)

            # Ask for save location and save
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="contracts_export.xlsx",
            )

            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Success", f"Exported to Excel:\n{file_path}")
        except (IOError, OSError, ValueError) as e:
            messagebox.showerror("Error", f"Excel export failed:\n{str(e)}")

    def _add_excel_headers(self, ws):
        """Add headers to Excel worksheet with styling"""
        headers = self.title_list
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _add_excel_data(self, ws):
        """Add treeview data to Excel worksheet"""
        for row_idx, item in enumerate(self.tree.get_children(), 2):
            item_data = self.tree.item(item)
            row_values = list(item_data.get("values", []))
            for col, value in enumerate(row_values, 1):
                ws.cell(row=row_idx, column=col, value=value)

    def _adjust_excel_columns(self, ws):
        """Adjust column widths in Excel worksheet"""
        for column in ws.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            except (AttributeError, TypeError):
                pass

    def apply_advanced_filters(self, filters: dict):
        def match(r: dict) -> bool:
            if filters.get("fullname"):
                if filters["fullname"].lower() not in str(r.get("fullname", "")).lower():
                    return False

            if filters.get("device_model"):
                if filters["device_model"].lower() not in str(r.get("Device_Model", "")).lower():
                    return False

            if filters.get("service_provider"):
                if filters["service_provider"].lower() not in str(r.get("ServiceMan", "")).lower():
                    return False

            if filters.get("problem_type"):
                if filters["problem_type"].lower() not in str(r.get("Device_Problem", "")).lower():
                    return False

            if filters.get("favorites_only") and not r.get("is_favorite"):
                return False

            if filters.get("date_from"):
                try:
                    if r.get("created_at") < filters["date_from"]:
                        return False
                except Exception:
                    pass

            if filters.get("date_to"):
                try:
                    if r.get("created_at") > filters["date_to"]:
                        return False
                except Exception:
                    pass

            return True

        # Determine if any meaningful filter is set
        meaningful = any(
            bool(v)
            for k, v in filters.items()
            if k not in ("date_from", "date_to") or v
        )

        if not meaningful:
            # No filters: reset to full dataset
            self._filters_active = False
            self.filtered_rows = self.rows.copy()
            self._render_tree(self.filtered_rows)
            return

        self.filtered_rows = [r for r in self.rows if match(r)]
        self._filters_active = True
        self._render_tree(self.filtered_rows)


    def _render_tree(self, rows: list[dict]):
        self.tree.delete(*self.tree.get_children())

        for r in rows:
            values = [r.get(col, "") for col in self.title_list]
            self.tree.insert("", "end", values=values)

    def populate_treeview(self):
        self.rows = self.db.get_all()
        self.filtered_rows = self.rows.copy()
        self._filters_active = False
        self._render_tree(self.filtered_rows)

    def search(self, query: str):
        # Normalize and tokenize query so multiple words (e.g. model + name)
        q = (query or "").strip().lower()

        # Clear selection and disable action buttons
        self.selected_record_id = None
        try:
            self.export_button.configure(state="disabled")
            self.view_details_button.configure(state="disabled")
            self.favorite_button.configure(state="disabled")
            self.print_button.configure(state="disabled")
            self.copy_button.configure(state="disabled")
        except Exception:
            pass

        # Ensure rows loaded
        if not self.rows:
            self.populate_treeview()

        # Decide base rows: if filters active, search within filtered_rows
        base_rows = self.filtered_rows if self._filters_active else self.rows

        if not q:
            # If query empty, just show base rows
            self._render_tree(base_rows)
            return

        tokens = [t for t in q.split() if t]
        search_columns = [
            "fullname",
            "Device_Model",
            "Device_Serial",
            "ServiceMan",
            "Device_Problem",
            "Description",
        ]

        def match_row_tokens(row: dict) -> bool:
            # Require that ALL tokens are found somewhere in the searchable columns (AND semantics)
            for token in tokens:
                found = False
                for col in search_columns:
                    try:
                        if token in str(row.get(col, "")).lower():
                            found = True
                            break
                    except Exception:
                        continue
                if not found:
                    return False
            return True

        results = [r for r in base_rows if match_row_tokens(r)]
        # Update filtered_rows to show search results (but keep filters_active state)
        self.filtered_rows = results
        self._render_tree(self.filtered_rows)


    def export_selected_contract(self):
        """Export contract for selected record"""
        if not self.selected_record_id:
            messagebox.showwarning("No Selection", "Please select a record to export.")
            return

        # Get record data
        try:
            record = self.db.get_by_id(int(self.selected_record_id))
        except Exception:
            record = None
        if not record:
            messagebox.showerror("Error", DATA_RETRIEVE_ERROR)
            return

        # Ensure data is in database (it should already be, but verify)
        # Check if record exists, if not, save it
        if not record.get("id"):
            # If somehow the record doesn't have an ID, save it to DB
            create_data = {
                "fullname": record.get("fullname", ""),
                "Device_Model": record.get("Device_Model", ""),
                "Device_Serial": record.get("Device_Serial", ""),
                "ServiceMan": record.get("ServiceMan", ""),
                "Device_Problem": record.get("Device_Problem", ""),
                "Description": record.get("Description", "")
                or record.get("description", ""),
            }
            try:
                inserted_id = self.db.create(create_data)
                if inserted_id:
                    record["id"] = inserted_id
                    messagebox.showinfo("Info", "Record saved to database.")
                else:
                    messagebox.showerror("Error", "Failed to save record to database.")
            except Exception as e:
                messagebox.showerror(
                    "Error", f"Failed to save record to database:\n{e}"
                )

        # Convert record to format expected by form handler
        # Map database column names to form field names
        # Handle both 'Description' (DB) and 'description' (form) for compatibility
        description_value = record.get("Description", "") or record.get(
            "description", ""
        )
        form_data = {
            "fullname": record.get("fullname", ""),
            "Device_Model": record.get("Device_Model", ""),
            "Device_Serial": record.get("Device_Serial", ""),
            "ServiceMan": record.get("ServiceMan", ""),
            "Device_Problem": record.get("Device_Problem", ""),
            "description": description_value,
            # Pass through the original saved timestamp so exported document
            # can show the actual save date instead of current time
            "created_at": record.get("created_at", ""),
        }

        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"Contract_{record.get('fullname', 'Unknown')}_{self.selected_record_id}.pdf",
        )

        if file_path:
            try:
                form_docx_to_pdf_handler([form_data], file_path)
                messagebox.showinfo(
                    "Success",
                    f"Contract exported successfully!\n\nFile saved to:\n{file_path}\n\nData is saved in database.",
                )
                # Refresh the treeview to show updated data
                self.populate_treeview()
            except (IOError, OSError, ValueError) as e:
                messagebox.showerror("Error", f"Failed to export contract:\n{str(e)}")

    def view_selected_details(self):
        """Show detailed view of selected record"""
        if not self.selected_record_id:
            messagebox.showwarning("No Selection", "Please select a record to view.")
            return

        # Get record data
        try:
            record = self.db.get_by_id(int(self.selected_record_id))
        except Exception:
            record = None
        if not record:
            messagebox.showerror("Error", DATA_RETRIEVE_ERROR)
            return

        # Create detail window
        detail_window = ctk.CTkToplevel(self.parent_window)
        detail_window.title("Record Details")
        detail_window.geometry("500x400")
        set_icon(detail_window)
        # Create scrollable text widget
        text_frame = ctk.CTkFrame(detail_window)
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)

        details_text = ctk.CTkTextbox(
            text_frame, height=300, font=ctk.CTkFont(family=PERSIAN_FONT, size=12)
        )
        details_text.pack(fill="both", expand=True, padx=5, pady=5)

        # Format and display details
        details_content = "RECORD DETAILS\n" + "=" * 50 + "\n\n"
        for key, value in record.items():
            if key != "id":  # Skip ID in display
                formatted_key = key.replace("_", " ").title()
                details_content += f"{formatted_key}:\n{value}\n\n"

        details_text.insert("1.0", details_content)
        details_text.configure(state="disabled")  # Make read-only

        # Close button
        close_button = ctk.CTkButton(
            detail_window, text="Close", command=detail_window.destroy
        )
        close_button.pack(pady=10)
